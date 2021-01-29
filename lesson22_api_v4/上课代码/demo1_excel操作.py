"""
excel操作的工具库
openyxl 主要针对xlsx格式的excel进行读取和编辑
xlrd 从excel中读取数据，支持 xls,xlsx
xlwt库：对excel进行修改操作，不支持对xlsx格式的修改
pandas：csv,数据分析
pip install openpyxl -i http://mirrors.aliyun.com/pypi/simple --trusted-host=mirrors.aliyun.com
"""
import pprint

import openpyxl

# 打开excel文件
from openpyxl.worksheet.worksheet import Worksheet

workbook = openpyxl.open('cases.xlsx')
print(workbook)

# 选择sheet表格，worksheet
# 通过表格的名称去获取：类似于字典的操作
ws : Worksheet = workbook['register']
print(ws)

# 获取数据，某个单元格的数据
cell = ws.cell(row=1,column=2)
print(cell.value)
print(ws.cell(row=3,column=5).value)

# 获取到第一行
print(ws[1])

#获取所有行，得到一个生成器,可以进行for循环，可以转换为list
print(ws.rows)
print(list(ws.rows))

#获取所有的数据
print(ws.values)
pprint.pprint(list(ws.values))
